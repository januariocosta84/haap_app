import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'haap_platform.settings')
django.setup()

from openpyxl import load_workbook
from core.models import Municipality, AdministrativePost, Suco, Aldeia

file_path = 'munic.xlsx'

workbook = load_workbook(file_path)
sheet = workbook.active  

for row in sheet.iter_rows(min_row=2, values_only=True):
    municipality_name = row[1]  # Column 2: Municipality
    administrative_post_name = row[2]  # Column 3: Administrative Post
    suco_name = row[3]  # Column 4: Suco
    village_name = row[4]  # Column 5: Village

    # Skip empty rows
    if not municipality_name:
        continue

    # Municipality
    municipality, _ = Municipality.objects.get_or_create(name=municipality_name)

    # Administrative Post
    administrative_post, _ = AdministrativePost.objects.get_or_create(
        name=administrative_post_name,
        municipality=municipality
    )

    # Suco
    suco, _ = Suco.objects.get_or_create(
        name=suco_name,
        administrative_post=administrative_post   # ✅ fixed field name
    )

    # Aldeia
    if village_name:  # allow empty if Excel has blanks
        aldeia, _ = Aldeia.objects.get_or_create(
            name=village_name,
            suco=suco
        )

    # Debug
    print(f"✅ Municipality: {municipality.name} | AP: {administrative_post.name} | Suco: {suco.name} | Aldeia: {village_name}")
print("Data import completed.")